#!/usr/bin/env python3
"""
Repara el estado «terminado» de `tasks` contra la verdad de la hoja.

Dos arreglos, con causas distintas, que se piden por separado:

**`--separador`**. En la hoja, «terminada» es una **posición**: las activas van
arriba, luego un hueco de filas vacías, y debajo la sección de tareas
realizadas. La base no guarda posiciones —lo calcula de la fila con
`_esta_archivada`— y esa fue una decisión deliberada de la migración. El
problema es que las dos verdades discrepan: hay 259 filas que su dueño bajó a
realizadas pero cuyo AVANCE se quedó en 10 % o 33 %, así que la aplicación las
sigue mostrando en operativo. Este modo cierra esas filas.

Que el hueco es señal fiable no es una suposición: de las 4.186 filas que están
debajo del separador en las 40 hojas, 3.927 (el 93,8 %) ya cumplen la regla por
su propio AVANCE o ESTATUS. Las 259 restantes son el desajuste.

**`--copias`**. Un folio de secuencia global (`PPC-`, `AV-`, `TG-`) tiene una
fila por hoja desde que se cargó el export en modo copia, y esas filas son la
misma actividad vista desde tablas distintas. `CAMPOS_BILATERALES` de
`api/services/asignacion.py` ya declara que AVANCE, ESTATUS y CUMPLIMIENTO
viajan entre copias; esto lo aplica a las que quedaron desincronizadas. Sin
esto, la fila del PPC maestro refleja su propia hoja —que trae ceros viejos
porque el cierre se captura en la pestaña de cada quien— y 959 actividades
terminadas reaparecen en operativo.

Lo que **no** hace, a propósito:

* No toca una fila donde la base está más avanzada que la hoja. Ahí la base es
  lo nuevo —la app escribe en Supabase, la hoja no— y pisarla borraría trabajo.
* No borra nada. Hay 584 filas en la base que la hoja ya no tiene porque
  alguien las quitó; se listan con `--huerfanas` y se decide aparte.

Por defecto no escribe: hay que pedir `--ejecutar`.

Uso:
    python scripts/sincronizar_terminadas.py --separador --archivo export.xlsx
    python scripts/sincronizar_terminadas.py --separador --archivo export.xlsx --ejecutar
    python scripts/sincronizar_terminadas.py --copias --ejecutar
    python scripts/sincronizar_terminadas.py --huerfanas --archivo export.xlsx
"""

from __future__ import annotations

import argparse
import collections
import json
import pathlib
import sys
import urllib.parse
import urllib.request
from typing import Any, Dict, List, Optional, Sequence, Tuple

RAIZ = pathlib.Path(__file__).resolve().parent.parent
sys.path.insert(0, str(RAIZ))

from api.services.asignacion import es_delegacion_de_fase  # noqa: E402
from api.services.sheets import _esta_archivada  # noqa: E402
from backend.core.config import cargar_settings  # noqa: E402
from backend.core.engines.postgrest import PostgrestEngine  # noqa: E402
from backend.schemas.hoja import clave_encabezado  # noqa: E402

LOTE = 200

# Columnas que definen el cierre. Es el subconjunto de `CAMPOS_BILATERALES` que
# decide el archivado; COMENTARIOS queda fuera porque el texto de una persona
# no es el de la otra.
CAMPOS_DE_CIERRE = ("avance", "status", "cumplimiento")

# NOT NULL sin DEFAULT: un upsert valida la tupla del INSERT **antes** de
# resolver el conflicto, así que hay que reenviarlas aunque no cambien.
# Es el mismo motivo que documenta `_completar_obligatorias`.
OBLIGATORIAS = ("folio", "concepto", "source_sheet")

# Filas vacías seguidas que cuentan como separador. Dos es poco —hay hojas con
# un renglón suelto en blanco entre bloques— y las reales van de 2 a 24, así
# que el corte se pone donde no hay ambigüedad.
HUECO_MINIMO = 3


def _cargar_env() -> None:
    """Vuelca el `.env` de la raíz en el entorno, sin pisar lo ya definido."""
    import os

    ruta = RAIZ / ".env"
    if not ruta.exists():
        return
    for linea in ruta.read_text().splitlines():
        linea = linea.strip()
        if not linea or linea.startswith("#") or "=" not in linea:
            continue
        clave, valor = linea.split("=", 1)
        os.environ.setdefault(clave.strip(), valor.strip())


def leer_tabla(url: str, key: str, tabla: str, columnas: Sequence[str]) -> List[Dict[str, Any]]:
    """
    Todas las filas de una tabla, en páginas de mil.

    No usa `PostgrestEngine.select` porque su firma no expone desplazamiento y
    PostgREST corta en 1.000 filas: `tasks` tiene siete mil.
    """
    cabeceras = {"apikey": key, "Authorization": f"Bearer {key}"}
    filas: List[Dict[str, Any]] = []
    pagina = 0
    while True:
        ruta = (f"{url.rstrip('/')}/rest/v1/{tabla}"
                f"?select={','.join(columnas)}&limit=1000&offset={pagina * 1000}")
        with urllib.request.urlopen(
                urllib.request.Request(ruta, headers=cabeceras), timeout=90) as respuesta:
            lote = json.load(respuesta)
        filas.extend(lote)
        if len(lote) < 1000:
            return filas
        pagina += 1


# ---------------------------------------------------------------------------
# La hoja: qué hay debajo del separador
# ---------------------------------------------------------------------------

ROTULO = "TAREAS REALIZADAS"


def _plano(valor: Any) -> str:
    """Mayúsculas, sin acentos y con los espacios colapsados."""
    import unicodedata

    if valor is None:
        return ""
    crudo = unicodedata.normalize("NFKD", str(valor).upper())
    return " ".join("".join(c for c in crudo if not unicodedata.combining(c)).split())


def fila_del_rotulo(ws: Any, fila_encabezado: int) -> Optional[int]:
    """
    Fila donde la hoja escribe «TAREAS REALIZADAS», o `None` si no lo escribe.

    Es el separador de verdad, y manda sobre el hueco. Dos cosas hacen que sea
    fácil no encontrarlo: no está en la columna del folio sino en la del
    concepto, y en `ADMINISTRADOR` está en la fila 1.864, así que cualquier
    búsqueda acotada a las primeras filas lo pierde. Lo escriben 73 de las 104
    hojas, incluidas las de ventas.
    """
    for fila in range(fila_encabezado, (ws.max_row or 0) + 1):
        for columna in range(1, min(ws.max_column or 0, 25) + 1):
            if _plano(ws.cell(fila, columna).value).startswith(ROTULO):
                return fila
    return None


def separar_por_hueco(ws: Any, fila_encabezado: int) -> Tuple[List[int], List[int]]:
    """
    `(filas_arriba, filas_abajo)` del separador de la hoja.

    El separador no es un rótulo sino un hueco: ninguna de las 40 hojas reales
    escribe «TAREAS REALIZADAS», simplemente dejan filas vacías entre el bloque
    activo y el histórico. Se corta en el primer hueco de `HUECO_MINIMO` filas.
    """
    con_datos = [r for r in range(fila_encabezado + 1, (ws.max_row or 0) + 1)
                 if ws.cell(r, 1).value not in (None, "")]
    if not con_datos:
        return [], []

    arriba: List[int] = []
    previa: Optional[int] = None
    for fila in con_datos:
        if previa is not None and fila - previa > HUECO_MINIMO:
            break
        arriba.append(fila)
        previa = fila

    corte = len(arriba)
    return arriba, con_datos[corte:]


def _folios_bajo(ruta: pathlib.Path) -> Dict[str, set]:
    """
    `{hoja: {folios de la sección de realizadas}}`.

    Reusa la detección de encabezado del cargador en vez de repetirla: las
    hojas reales lo tienen en la fila 9, la 10 o la 12, y mantener dos copias
    de esa heurística garantiza que se desincronicen.
    """
    import importlib.util

    import openpyxl

    spec = importlib.util.spec_from_file_location(
        "cargar_tracker_xlsx", RAIZ / "scripts" / "cargar_tracker_xlsx.py")
    cargador = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(cargador)

    libro = openpyxl.load_workbook(ruta, data_only=True)
    resultado: Dict[str, set] = {}
    for nombre in libro.sheetnames:
        hoja = libro[nombre]
        if not hoja.max_row or hoja.max_row < 2:
            continue
        ubicacion = cargador.localizar_encabezado(hoja)
        if ubicacion is None:
            continue
        abajo = filas_de_realizadas(hoja, ubicacion[0])
        if not abajo:
            continue
        encabezados = ubicacion[1]
        col_avance = next((i for i, e in enumerate(encabezados, start=1)
                           if clave_encabezado(e).startswith("AVANCE")), None)
        resultado[nombre] = {
            str(hoja.cell(r, 1).value).strip():
                (hoja.cell(r, col_avance).value if col_avance else None)
            for r in abajo if hoja.cell(r, 1).value not in (None, "")
        }
    return resultado


def filas_de_realizadas(ws: Any, fila_encabezado: int) -> List[int]:
    """
    Filas de la sección de realizadas: manda el rótulo, y solo si falta el hueco.

    El rótulo es explícito y el hueco es una inferencia, así que no compiten:
    donde hay rótulo, el hueco se ignora. La diferencia no es teórica —en
    `SONIA GARCIA PEREZ` el hueco deja 247 filas arriba y el rótulo solo 29—,
    y las 31 hojas sin rótulo siguen necesitando el hueco para no quedarse sin
    sección de realizadas.
    """
    rotulo = fila_del_rotulo(ws, fila_encabezado)
    if rotulo is None:
        return separar_por_hueco(ws, fila_encabezado)[1]
    return [fila for fila in range(rotulo + 1, (ws.max_row or 0) + 1)
            if ws.cell(fila, 1).value not in (None, "")]


# ---------------------------------------------------------------------------
# Los dos arreglos
# ---------------------------------------------------------------------------

def _norm(valor: Any) -> str:
    return " ".join(str(valor or "").split()).upper()


# Firma de una hoja de ventas: comparte FOLIO y AVANCE con las de tareas, pero
# CLIENTE y COTIZACION solo existen aquí.
FIRMA_VENTAS = frozenset({"FOLIO", "AVANCE", "ESTATUS", "CLIENTE"})


def folios_realizados_de_ventas(ruta: pathlib.Path) -> Dict[str, set]:
    """
    `{hoja de ventas: {folios bajo el rótulo}}`.

    Las trece hojas `(VENTAS)` usan el mismo separador que los trackers
    —`ANTONIA_VENTAS` lo tiene en la fila 192— pero alimentan `quotes`, con
    otro esquema y otra clave (`folio` + `source_sheet`), así que no pueden
    pasar por el mismo camino que las tareas.
    """
    import openpyxl

    libro = openpyxl.load_workbook(ruta, data_only=True)
    resultado: Dict[str, set] = {}
    for nombre in libro.sheetnames:
        hoja = libro[nombre]
        if not hoja.max_row or hoja.max_row < 2:
            continue
        encabezado = None
        for fila in range(1, min(hoja.max_row, 12) + 1):
            claves = {clave_encabezado(hoja.cell(fila, c).value)
                      for c in range(1, (hoja.max_column or 0) + 1)}
            if FIRMA_VENTAS <= claves:
                encabezado = fila
                break
        if encabezado is None:
            continue
        rotulo = fila_del_rotulo(hoja, encabezado)
        if rotulo is None:
            continue
        columnas = {clave_encabezado(hoja.cell(encabezado, c).value): c
                    for c in range(1, (hoja.max_column or 0) + 1)}
        folios: Dict[str, Dict[str, Any]] = {}
        for fila in range(rotulo + 1, (hoja.max_row or 0) + 1):
            folio = hoja.cell(fila, 1).value
            if folio in (None, ""):
                continue
            folios[str(folio).strip()] = {
                "avance": hoja.cell(fila, columnas["AVANCE"]).value if "AVANCE" in columnas else None,
                "estatus": hoja.cell(fila, columnas["ESTATUS"]).value if "ESTATUS" in columnas else None,
            }
        if folios:
            resultado[nombre] = folios
    return resultado


def cotizaciones_abiertas_bajo_el_rotulo(
    filas: Sequence[Dict[str, Any]], bajo: Dict[str, set]
) -> List[Dict[str, Any]]:
    """Cotizaciones que la hoja tiene como realizadas y la base sigue abriendo."""
    from api.services.tracker_rules import is_progress_complete_pct, is_terminal_status

    def cerrada(fila: Dict[str, Any]) -> bool:
        try:
            avance = float(fila.get("avance")) if fila.get("avance") is not None else None
        except (TypeError, ValueError):
            avance = None
        if avance is not None and is_progress_complete_pct(avance):
            return True
        return is_terminal_status(fila.get("estatus"))

    indice = {_norm(h): f for h, f in bajo.items()}
    pendientes: List[Dict[str, Any]] = []
    for fila in filas:
        hoja = indice.get(_norm(fila.get("source_sheet")))
        if not hoja:
            continue
        dato = hoja.get(str(fila.get("folio") or "").strip())
        if dato is None or cerrada(fila):
            continue
        # El estatus de la hoja es lo que cierra: media docena son
        # «Perdida x Tiempo» o «Cancelada x Planta», y marcarlas al 100 %
        # diría que se entregaron cuando lo que pasó es que se perdieron.
        pendientes.append({**fila, "_hoja": dato})
    return pendientes


def desajustes_por_separador(
    filas: Sequence[Dict[str, Any]], bajo: Dict[str, set]
) -> List[Dict[str, Any]]:
    """Filas que la hoja tiene como realizadas y la base sigue mostrando activas."""
    indice = {_norm(h): f for h, f in bajo.items()}
    pendientes: List[Dict[str, Any]] = []
    for fila in filas:
        folios = indice.get(_norm(fila.get("source_sheet")))
        if folios is None:
            continue
        folio = str(fila.get("folio") or "").strip()
        if folio not in folios:
            continue
        # Solo importa si la base la sigue mostrando activa. No se compara el
        # avance contra el de la hoja: estar debajo del rótulo la cierra al
        # 100 % por decisión del dueño del dato, así que la hoja diría 33 % para
        # siempre y la fila se re-detectaría en cada pasada sin converger nunca.
        if not _esta_archivada(fila):
            pendientes.append({**fila, "_avance_hoja": folios[folio]})
    return pendientes


def _avance(fila: Dict[str, Any]) -> float:
    try:
        return float(fila.get("avance") or 0)
    except (TypeError, ValueError):
        return 0.0


def copias_desincronizadas(
    filas: Sequence[Dict[str, Any]],
) -> List[Tuple[Dict[str, Any], Dict[str, Any]]]:
    """
    `(copia_abierta, copia_cerrada)` por cada copia que quedó atrás.

    Se agrupa por folio y no por `dedupe_key`, que es justamente lo que difiere
    entre copias. Solo se propaga cuando existe una hermana cerrada: si todas
    están abiertas no hay nada que propagar y cerrarlas sería inventar.

    Las microtareas de papa caliente se excluyen. No es cautela genérica:
    comparten folio entre sí y con la cotización, así que propagar pondría el
    100 % de una fase encima del trabajo pendiente de otra persona. Es la misma
    frontera que traza `campos_sincronizables`.
    """
    por_folio: Dict[str, List[Dict[str, Any]]] = collections.defaultdict(list)
    for fila in filas:
        folio = str(fila.get("folio") or "").strip()
        if folio:
            por_folio[folio].append(fila)

    pendientes: List[Tuple[Dict[str, Any], Dict[str, Any]]] = []
    for hermanas in por_folio.values():
        if len(hermanas) < 2:
            continue
        cerradas = [f for f in hermanas if _esta_archivada(f)]
        if not cerradas:
            continue
        modelo = max(cerradas, key=_avance)
        for fila in hermanas:
            if _esta_archivada(fila):
                continue
            if es_delegacion_de_fase({"CONCEPTO": fila.get("concepto") or ""}):
                continue
            pendientes.append((fila, modelo))
    return pendientes


def fila_de_cierre(fila: Dict[str, Any], modelo: Optional[Dict[str, Any]] = None) -> Dict[str, Any]:
    """
    Carga útil del upsert: la identidad, las obligatorias y el cierre.

    Sin `modelo` se cierra con avance 100 y `CUMPLIMIENTO = SI`.

    Las dos cosas, y el 100 es una decisión del dueño del dato, no una
    inferencia: estar debajo del rótulo significa terminada, y un avance de
    10 % o 33 % ahí abajo es residuo de la plataforma anterior, no progreso
    real. `CUMPLIMIENTO` se pone igualmente porque es la bandera que consulta
    `_esta_archivada` y no depende de cómo se lea el número.

    Con `modelo` se copia el estado de la copia hermana, sin inventar valores.
    """
    carga: Dict[str, Any] = {"dedupe_key": fila["dedupe_key"]}
    for columna in OBLIGATORIAS:
        carga[columna] = fila.get(columna)

    if modelo is None:
        carga["avance"] = 100
        carga["cumplimiento"] = "SI"
        return carga

    for campo in CAMPOS_DE_CIERRE:
        valor = modelo.get(campo)
        if valor is not None and str(valor).strip() != "":
            carga[campo] = valor
    return carga


def _cierre_de_cotizacion(fila: Dict[str, Any]) -> Dict[str, Any]:
    """
    Carga útil para cerrar una cotización, con el estado que trae la hoja.

    Misma regla que en `tasks`: debajo del rótulo es cerrada, con avance 100.

    `completada` es la bandera equivalente a `cumplimiento` —lo que consulta
    `_esta_archivada` con `campo_cumplimiento="completada"`— pero **es boolean
    en el esquema, no texto**: mandarle `"SI"` aborta el lote con un 22P02
    («invalid input syntax for type boolean»). Comprobado contra la base.

    Si la hoja trae un estatus terminal se guarda además, porque dice *por qué*
    se cerró: «Perdida x Tiempo» y «Cancelada x Planta» no son lo mismo que
    entregada, y esa distinción es la que alimenta los KPIs de ventas.
    """
    from api.services.tracker_rules import is_terminal_status

    hoja = fila["_hoja"]
    carga: Dict[str, Any] = {"folio": fila["folio"], "source_sheet": fila["source_sheet"],
                             "avance": 100, "completada": True}
    estatus = hoja.get("estatus")
    if estatus not in (None, "") and is_terminal_status(estatus):
        carga["estatus"] = str(estatus).strip()
    return carga


def escribir(motor: Any, cargas: Sequence[Dict[str, Any]]) -> int:
    """Upsert por lotes, agrupando por conjunto de columnas."""
    escritas = 0
    grupos: Dict[tuple, List[Dict[str, Any]]] = collections.defaultdict(list)
    for carga in cargas:
        grupos[tuple(sorted(carga))].append(carga)
    for grupo in grupos.values():
        for inicio in range(0, len(grupo), LOTE):
            lote = grupo[inicio : inicio + LOTE]
            motor.upsert("tasks", lote, en_conflicto="dedupe_key")
            escritas += len(lote)
    return escritas


def _resumen(titulo: str, filas: Sequence[Dict[str, Any]]) -> None:
    print(f"\n{titulo}: {len(filas)}")
    for hoja, n in collections.Counter(f["source_sheet"] for f in filas).most_common(12):
        print(f"   {str(hoja)[:38]:40} {n}")


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--archivo", default=None, help="Export .xlsx del Tracker")
    parser.add_argument("--separador", action="store_true",
                        help="Cierra lo que la hoja tiene debajo del separador.")
    parser.add_argument("--copias", action="store_true",
                        help="Propaga el cierre entre copias del mismo folio.")
    parser.add_argument("--cotizaciones", action="store_true",
                        help="Cierra en `quotes` lo que la hoja de ventas tiene bajo el rótulo.")
    parser.add_argument("--huerfanas", action="store_true",
                        help="Cierra las filas que la hoja ya no tiene (con --ejecutar; sin él, lista).")
    parser.add_argument("--ejecutar", action="store_true",
                        help="Escribe de verdad. Sin esta bandera solo se cuenta.")
    args = parser.parse_args()

    if not (args.separador or args.copias or args.huerfanas or args.cotizaciones):
        parser.error("Indica al menos un modo: --separador, --copias, --cotizaciones o --huerfanas")
    if (args.separador or args.huerfanas or args.cotizaciones) and not args.archivo:
        parser.error("--separador, --cotizaciones y --huerfanas necesitan --archivo")

    _cargar_env()
    settings = cargar_settings()
    if not settings.hay_postgrest:
        print("Faltan SUPABASE_URL / SUPABASE_KEY.", file=sys.stderr)
        return 2
    motor = PostgrestEngine(settings.supabase_url, settings.supabase_key)

    filas = leer_tabla(settings.supabase_url, settings.supabase_key, "tasks",
                       ["dedupe_key", "folio", "source_sheet", "avance",
                        "status", "cumplimiento", "concepto"])
    activas = [f for f in filas if not _esta_archivada(f)]
    print(f"`tasks`: {len(filas)} filas | activas {len(activas)} | realizadas {len(filas) - len(activas)}")

    cargas: List[Dict[str, Any]] = []

    if args.separador:
        bajo = _folios_bajo(pathlib.Path(args.archivo))
        pendientes = desajustes_por_separador(filas, bajo)
        _resumen("Debajo del separador en la hoja pero ACTIVAS en la base", pendientes)
        for fila in pendientes[:6]:
            print(f"     {fila['folio']:14} {str(fila['source_sheet'])[:24]:26} "
                  f"avance {fila['avance']} -> 100")
        cargas.extend(fila_de_cierre(f) for f in pendientes)

    if args.copias:
        pendientes = copias_desincronizadas(filas)
        _resumen("Copias abiertas con una hermana ya cerrada", [f for f, _ in pendientes])
        for fila, modelo in pendientes[:6]:
            print(f"     {fila['folio']:14} {str(fila['source_sheet'])[:24]:26} "
                  f"avance {fila['avance']} -> {modelo['avance']} "
                  f"(de {str(modelo['source_sheet'])[:20]})")
        cargas.extend(fila_de_cierre(f, m) for f, m in pendientes)

    if args.cotizaciones:
        bajo = folios_realizados_de_ventas(pathlib.Path(args.archivo))
        cotizaciones = leer_tabla(settings.supabase_url, settings.supabase_key, "quotes",
                                  ["folio", "source_sheet", "avance", "estatus"])
        pendientes = cotizaciones_abiertas_bajo_el_rotulo(cotizaciones, bajo)
        print(f"\n`quotes`: {len(cotizaciones)} filas | "
              f"{len(bajo)} hojas de ventas con rótulo")
        _resumen("Bajo el rótulo en la hoja pero ABIERTAS en la base", pendientes)
        for fila in pendientes[:8]:
            print(f"     {fila['folio']:12} {str(fila['source_sheet'])[:24]:26} "
                  f"bd({fila['avance']}, {str(fila['estatus'])[:18]!r}) -> "
                  f"hoja({fila['_hoja']['avance']}, {str(fila['_hoja']['estatus'])[:18]!r})")
        if args.ejecutar and pendientes:
            # Agrupadas por conjunto de columnas: PostgREST exige que todos los
            # objetos de un mismo arreglo tengan las mismas claves, y `estatus`
            # solo viaja en las que la hoja cierra con un motivo.
            escritas = 0
            grupos: Dict[tuple, List[Dict[str, Any]]] = collections.defaultdict(list)
            for fila in pendientes:
                carga = _cierre_de_cotizacion(fila)
                grupos[tuple(sorted(carga))].append(carga)
            for grupo in grupos.values():
                for inicio in range(0, len(grupo), LOTE):
                    motor.upsert("quotes", grupo[inicio : inicio + LOTE],
                                 en_conflicto="folio,source_sheet")
                    escritas += len(grupo[inicio : inicio + LOTE])
            print(f"   ESCRITO: {escritas} cotizaciones cerradas.")

    if args.huerfanas:
        import importlib.util

        spec = importlib.util.spec_from_file_location(
            "cargar_tracker_xlsx", RAIZ / "scripts" / "cargar_tracker_xlsx.py")
        cargador = importlib.util.module_from_spec(spec)
        spec.loader.exec_module(cargador)
        hojas = cargador.leer_hojas(pathlib.Path(args.archivo))
        en_hoja: Dict[str, set] = {}
        for nombre, contenido in hojas.items():
            tareas, _ = cargador.convertir(contenido)
            en_hoja[_norm(nombre)] = {str(t.folio).strip() for t in tareas}
        huerfanas = [f for f in filas
                     if _norm(f["source_sheet"]) in en_hoja
                     and str(f["folio"]).strip() not in en_hoja[_norm(f["source_sheet"])]]
        abiertas = [f for f in huerfanas if not _esta_archivada(f)]
        _resumen("En la base pero ya no en la hoja", huerfanas)
        print(f"\n   de ellas, activas hoy: {len(abiertas)}")
        # Se cierran solo las abiertas, y solo si se pide. Que una fila ya no
        # esté en la hoja no dice por sí solo que esté terminada: alguien la
        # quitó, y el motivo —hecha, duplicada, error de captura— no está en el
        # dato. Cerrarlas es la lectura que eligió el dueño; borrarlas no, y por
        # eso este script no borra nunca.
        cargas.extend(fila_de_cierre(f) for f in abiertas)

    if not cargas:
        print("\nNada que escribir.")
        return 0

    # Una fila puede caer en los dos arreglos; se manda una sola vez.
    unicas: Dict[str, Dict[str, Any]] = {}
    for carga in cargas:
        unicas.setdefault(carga["dedupe_key"], carga).update(carga)
    cargas = list(unicas.values())

    print(f"\nFilas a actualizar (sin repetir): {len(cargas)}")
    if args.ejecutar:
        print(f"ESCRITO EN LA BASE: {escribir(motor, cargas)} filas.")
    else:
        print("DRY-RUN: no se escribió nada.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
