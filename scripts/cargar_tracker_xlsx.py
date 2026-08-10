#!/usr/bin/env python3
"""
Carga un export .xlsx del Tracker sobre la base real, tabla `tasks`.

Existe porque el sync hoja -> Supabase vive en Google Apps Script
(`SupabaseSync.mirrorBatch`, `CODIGO.js:4006`) y solo se dispara cuando alguien
guarda desde la interfaz. Cuando el corte de la base se queda atrás respecto de
la hoja —el caso que motivó este script: base al 27 de julio, hoja al 8 de
agosto— no había forma de recuperar el rezago sin abrir y volver a guardar cada
tracker a mano.

**No reimplementa la escritura.** Todo el camino de guardado es el del
repositorio (`TasksRepository.guardar_lote`), que ya resuelve `dedupe_key`,
`source_sheet`, la normalización de `status`, las columnas NOT NULL y el
`assignee_id`. Este archivo solo hace tres cosas que el repositorio no hace:
leer el .xlsx, decidir qué hoja es de tareas, y contar antes de escribir.

Por defecto **no escribe**: hay que pedir `--ejecutar` de forma explícita.

Uso:
    python scripts/cargar_tracker_xlsx.py --archivo export.xlsx            # dry-run
    python scripts/cargar_tracker_xlsx.py --archivo export.xlsx --ejecutar
    python scripts/cargar_tracker_xlsx.py --archivo export.xlsx --hoja "TERESA GARZA"

Requiere SUPABASE_URL y SUPABASE_KEY (lee el `.env` de la raíz si existe).
"""

from __future__ import annotations

import argparse
import collections
import pathlib
import sys
from datetime import date, datetime, time
from typing import Any, Dict, Iterator, List, Optional, Sequence, Set, Tuple

RAIZ = pathlib.Path(__file__).resolve().parent.parent
sys.path.insert(0, str(RAIZ))

from backend.core.config import cargar_settings  # noqa: E402
from backend.core.engines.postgrest import PostgrestEngine  # noqa: E402
from backend.repositories.tasks import TaskRepository  # noqa: E402
from backend.schemas.hoja import clave_encabezado, numero_a_texto  # noqa: E402
from backend.schemas.task import (  # noqa: E402
    INDICE_ALIAS,
    OBLIGATORIAS_AL_INSERTAR,
    TIPOS_REALES,
    TaskWrite,
)
from backend.services.identity import compute_dedupe_key  # noqa: E402

# Firma mínima de una hoja de tareas. `AVANCE` y `STATUS` son lo que la
# distingue de una hoja de ventas, que también tiene `FOLIO` y `CONCEPTO` pero
# alimenta `quotes` con otro esquema (`CLIENTE`, `COTIZACION`, `PRIO. COT.`).
FIRMA_TAREAS = frozenset({"FOLIO", "AVANCE", "STATUS"})

# Un encabezado de ventas nunca debe entrar por aquí aunque cumpla la firma.
FIRMA_VENTAS = frozenset({"CLIENTE", "COTIZACION"})

# Hasta dónde se busca la fila de encabezado. Las hojas reales la tienen en la
# 9, la 10 o la 12 según cuánto encabezado decorativo arrastre cada tracker.
MAX_FILA_ENCABEZADO = 30

# `RELOJ` se descarta a propósito, y conviene dejar dicho por qué.
#
# En la base es `text` y su valor real es `"00:00:00"` en 3.923 de las 4.554
# filas; el resto son números sueltos ("283.0") y alguna hora ("17:00:00"). En
# el .xlsx la misma columna sale como **serial de fecha** de Excel: la celda de
# TG-0027 llega como `datetime(1900, 1, 7)`, o sea "7 días", mientras que en la
# base esa misma fila dice `"00:00:00"`. Son dos codificaciones distintas del
# mismo campo, y ninguna de las dos es la otra.
#
# Escribirla introduciría una tercera codificación en una columna que de todos
# modos es derivada: `calculateDiasCounter` la recalcula en cada carga del
# tracker. Omitirla es lo único que no pierde nada, porque el upsert va con
# `merge-duplicates` y una columna ausente conserva lo que ya había.
# `HORA ESTIMADA DE FIN` se queda fuera por una razón distinta, y conviene no
# confundirla con la de arriba: no es que su valor esté mal, es que hoy no hay
# forma de escribirlo sin cambiar código compartido.
#
# `ALIAS_DE_HOJA` declara `HORA ESTIMADA FIN`, `HORA_ESTIMADA_FIN` y
# `HR. EST. FIN`, pero **no** `HORA ESTIMADA DE FIN`, que es como se rotula la
# columna en las 40 hojas del export y como la lista `index.html:9018` entre
# sus columnas de hora. Sin ese alias, `columnas_desde_hoja` no la reconoce y
# la descarta: son 1.523 valores del export.
#
# La consecuencia ya está en la base —`hora_estimada_fin` tiene 838 filas con
# dato, todas de la migración inicial, ninguna escrita por este camino— así que
# el hueco es anterior a este script y afecta también al guardado del frontend.
# Añadir el alias lo arreglaría, pero es un cambio en un módulo compartido que
# toca `/api/legacy/saveTrackerBatch`, y no es algo que deba colarse dentro de
# una carga de datos. Se deja anotado aquí y fuera del lote: omitida, el
# `merge-duplicates` conserva las 838 filas que ya hay.
ENCABEZADOS_IGNORADOS = frozenset({"RELOJ", "HORAESTIMADADEFIN"})


def _a_texto_generico(encabezado: str, valor: Any, indice: Dict[str, str],
                      tipos: Dict[str, str]) -> Any:
    """
    `_a_texto_si_toca` parametrizado por el esquema de la tabla destino.

    La patología es la misma en tareas y en ventas —Excel tipa la celda y la
    base tipa la columna— pero los mapas de alias y de tipos son distintos, así
    que la regla se comparte y los diccionarios no.
    """
    columna = indice.get(clave_encabezado(encabezado))
    if not columna or tipos.get(columna) != "text":
        return valor
    if isinstance(valor, (datetime, date, time)):
        return str(valor)
    return numero_a_texto(valor)


def _a_texto_si_toca(encabezado: str, valor: Any) -> Any:
    """
    Un número en una columna `text` se guarda como texto; lo demás no se toca.

    Excel tipa la celda, no la columna. En una columna `text` de la base eso
    deja tres formas de la misma cosa: el folio timestamp `1772639658256` llega
    como `float`, y `restricciones` con solo un número dentro también. Pydantic
    los rechaza —son 327 filas del export— y perderlas por el tipo de la celda
    sería tirar dato bueno.

    La conversión la hace `numero_a_texto`, el mismo helper que usa el camino
    del frontend, y se aplica **solo** donde la columna destino es `text`: en
    `avance`, en las fechas y en las horas el tipo del valor es justamente lo
    que sus validadores necesitan para decidir.

    Las fechas en columna de texto van por `str()`, y el formato no es una
    elección: `restricciones` es texto libre y hay quien escribe una fecha
    dentro. La base ya guarda catorce de esas filas como
    `"2026-07-17 00:00:00"`, que es justo lo que `str(datetime)` produce, así
    que esto reproduce lo que el camino de Apps Script venía escribiendo en vez
    de introducir una grafía nueva.
    """
    return _a_texto_generico(encabezado, valor, INDICE_ALIAS, TIPOS_REALES)


# --- hojas de ventas -> `quotes` -------------------------------------------
#
# Se cargan por separado de las de tareas y con su propio repositorio: la clave
# de identidad es `(folio, source_sheet)` y no `dedupe_key`, el esquema tiene
# 37 columnas distintas, y una cotización no puede acabar en `tasks` —eso es lo
# que `FIRMA_VENTAS` impide en `localizar_encabezado`—.
#
# `RELOJ` se ignora por el mismo motivo que en tareas: el .xlsx lo entrega como
# serial de fecha. `COMPLETADA` también, pero por otro: es boolean en la base y
# la hoja no la escribe, así que dejarla pasar solo puede estropear el archivado
# que ya se sincroniza con `scripts/sincronizar_terminadas.py --cotizaciones`.
ENCABEZADOS_IGNORADOS_VENTAS = frozenset({"RELOJ", "COMPLETADA"})


# Firma de una hoja de ventas. `ESTATUS` y no `STATUS`: es la diferencia de
# rótulo entre los dos dominios y la razón por la que `FIRMA_TAREAS` no sirve
# aquí. `CLIENTE` es lo que ninguna hoja de tareas tiene.
FIRMA_HOJA_VENTAS = frozenset({"FOLIO", "CLIENTE", "ESTATUS"})


def _es_hoja_de_otra_tabla(nombre: str) -> bool:
    """
    ¿Esta hoja alimenta una tabla que no es `quotes`?

    `DB_BANCO_DATOS` cumple la firma de ventas —tiene FOLIO, CLIENTE y
    ESTATUS— pero `TABLAS_POR_HOJA` la enruta a `banco_datos`. Meterla en
    `quotes` inventaría tres cotizaciones que nadie cotizó. La lista de destinos
    se consulta en vez de repetirse para que añadir una hoja allí baste.

    Se prueba también sin el prefijo `DB_`: el mapa registra `BANCO_DATOS` y la
    pestaña se llama `DB_BANCO_DATOS`, así que la comparación literal la deja
    pasar.
    """
    from api.services.sheets import TABLAS_POR_HOJA

    limpio = str(nombre).strip().upper()
    return limpio in TABLAS_POR_HOJA or limpio.removeprefix("DB_") in TABLAS_POR_HOJA


def localizar_encabezado_ventas(ws: Any) -> Optional[Tuple[int, List[str]]]:
    """`(fila, encabezados)` de una hoja de ventas, o `None` si no lo es."""
    tope = min(ws.max_row or 0, MAX_FILA_ENCABEZADO)
    for fila in range(1, tope + 1):
        crudos = [ws.cell(fila, col).value for col in range(1, (ws.max_column or 0) + 1)]
        claves = {clave_encabezado(v) for v in crudos if v is not None}
        if FIRMA_HOJA_VENTAS <= claves:
            return fila, [("" if v is None else str(v)) for v in crudos]
    return None


def leer_hojas_de_ventas(ruta: pathlib.Path,
                         solo: Optional[str] = None) -> Dict[str, List[Dict[str, Any]]]:
    """`{hoja de ventas: [filas]}`, con los tipos ya conciliados."""
    import openpyxl

    from backend.schemas.quote import INDICE_ALIAS as ALIAS_Q
    from backend.schemas.quote import TIPOS_REALES as TIPOS_Q

    libro = openpyxl.load_workbook(ruta, data_only=True)
    resultado: Dict[str, List[Dict[str, Any]]] = {}
    for nombre in libro.sheetnames:
        if solo is not None and nombre != solo:
            continue
        if _es_hoja_de_otra_tabla(nombre):
            continue
        hoja = libro[nombre]
        if not hoja.max_row or hoja.max_row < 2:
            continue
        ubicacion = localizar_encabezado_ventas(hoja)
        if ubicacion is None:
            continue
        fila_encabezado, encabezados = ubicacion
        filas: List[Dict[str, Any]] = []
        for numero in range(fila_encabezado + 1, (hoja.max_row or 0) + 1):
            folio = hoja.cell(numero, 1).value
            if folio is None or str(folio).strip() == "":
                continue
            fila = {
                encabezado: _a_texto_generico(encabezado, hoja.cell(numero, col).value,
                                              ALIAS_Q, TIPOS_Q)
                for col, encabezado in enumerate(encabezados, start=1)
                if encabezado and clave_encabezado(encabezado) not in ENCABEZADOS_IGNORADOS_VENTAS
            }
            fila["_fila_xlsx"] = numero
            filas.append(fila)
        if filas:
            resultado[nombre] = filas
    return resultado


def convertir_ventas(filas: Sequence[Dict[str, Any]]) -> Tuple[List[Any], List[Tuple[int, str]]]:
    """`[QuoteWrite]` y los rechazos, con la misma tolerancia que en tareas."""
    from backend.schemas.quote import QuoteWrite

    convertidas: List[Any] = []
    rechazos: List[Tuple[int, str]] = []
    for fila in filas:
        numero = fila.pop("_fila_xlsx", -1)
        try:
            convertidas.append(QuoteWrite.desde_hoja(fila))
        except Exception as exc:  # noqa: BLE001 - se reporta, no se traga
            rechazos.append((numero, str(exc).replace("\n", " ")[:160]))
    return convertidas, rechazos

LOTE = 200

# Claves por consulta. Van dentro de la URL, así que el tope no es de memoria
# sino de longitud de petición.
LOTE_CONSULTA = 100


def _cargar_env() -> None:
    """Vuelca el `.env` de la raíz en el entorno, sin pisar lo ya definido."""
    import os

    ruta = RAIZ / ".env"
    if not ruta.exists():
        return
    for linea in ruta.read_text().splitlines():
        linea = linea.strip()
        if not linea or linea.startswith("#") or "=" not in linea:
            continue
        clave, valor = linea.split("=", 1)
        os.environ.setdefault(clave.strip(), valor.strip())


def localizar_encabezado(ws: Any) -> Optional[Tuple[int, List[str]]]:
    """
    `(fila, encabezados)` de la hoja, o `None` si no es una hoja de tareas.

    Se busca por contenido y no por posición fija porque las 44 hojas de
    tareas reales no coinciden: `ADMINISTRADOR` la tiene en la fila 9,
    la mayoría en la 10 y `ANTONIO SALAZAR` en la 12.
    """
    tope = min(ws.max_row or 0, MAX_FILA_ENCABEZADO)
    for fila in range(1, tope + 1):
        crudos = [ws.cell(fila, col).value for col in range(1, (ws.max_column or 0) + 1)]
        claves = {clave_encabezado(v) for v in crudos if v is not None}
        if FIRMA_TAREAS <= claves and not (FIRMA_VENTAS & claves):
            return fila, [("" if v is None else str(v)) for v in crudos]
    return None


def filas_de_hoja(ws: Any, fila_encabezado: int, encabezados: Sequence[str]) -> Iterator[Dict[str, Any]]:
    """
    Cada fila con folio, como dict `{encabezado: valor}`.

    Se descarta la fila sin folio: en estas hojas las filas vacías por debajo
    del último dato llegan hasta la 1000 y pico, y una fila sin folio no tiene
    `dedupe_key` posible.
    """
    for numero in range(fila_encabezado + 1, (ws.max_row or 0) + 1):
        folio = ws.cell(numero, 1).value
        if folio is None or str(folio).strip() == "":
            continue
        fila = {
            encabezado: _a_texto_si_toca(encabezado, ws.cell(numero, col).value)
            for col, encabezado in enumerate(encabezados, start=1)
            if encabezado and clave_encabezado(encabezado) not in ENCABEZADOS_IGNORADOS
        }
        fila["_fila_xlsx"] = numero
        yield fila


def leer_hojas(ruta: pathlib.Path, solo: Optional[str] = None) -> Dict[str, List[Dict[str, Any]]]:
    """`{nombre_de_hoja: [filas]}` para las hojas con firma de tareas."""
    import openpyxl

    libro = openpyxl.load_workbook(ruta, data_only=True)
    resultado: Dict[str, List[Dict[str, Any]]] = {}
    for nombre in libro.sheetnames:
        if solo is not None and nombre != solo:
            continue
        hoja = libro[nombre]
        if not hoja.max_row or hoja.max_row < 2:
            continue
        ubicacion = localizar_encabezado(hoja)
        if ubicacion is None:
            continue
        fila_encabezado, encabezados = ubicacion
        filas = list(filas_de_hoja(hoja, fila_encabezado, encabezados))
        if filas:
            resultado[nombre] = filas
    return resultado


def convertir(filas: Sequence[Dict[str, Any]]) -> Tuple[List[TaskWrite], List[Tuple[int, str]]]:
    """
    `[TaskWrite]` y los rechazos como `(fila_xlsx, motivo)`.

    Una fila que no valida no aborta la hoja: se aparta y se reporta. Un export
    manual trae erratas puntuales y perder las 300 filas buenas de un tracker
    por una celda mal escrita sería peor que dejar esa celda fuera.
    """
    convertidas: List[TaskWrite] = []
    rechazos: List[Tuple[int, str]] = []
    for fila in filas:
        numero = fila.pop("_fila_xlsx", -1)
        try:
            convertidas.append(TaskWrite.desde_hoja(fila))
        except Exception as exc:  # noqa: BLE001 - se reporta, no se traga
            rechazos.append((numero, str(exc).replace("\n", " ")[:160]))
    return convertidas, rechazos


def falta_para_el_alta(tarea: TaskWrite) -> List[str]:
    """
    Columnas obligatorias que esta fila **nueva** no trae.

    `tasks` tiene cuatro columnas NOT NULL sin DEFAULT, y de ellas `dedupe_key`
    y `source_sheet` las pone el repositorio: al export solo le tocan `folio` y
    `concepto`. En una actualización esto no aplica, porque
    `_completar_obligatorias` las toma de la fila que ya está en la base.
    """
    return sorted(
        columna
        for columna in OBLIGATORIAS_AL_INSERTAR & {"folio", "concepto"}
        if not str(getattr(tarea, columna, "") or "").strip()
    )


def colapsar_claves_repetidas(
    hoja: str, tareas: Sequence[TaskWrite]
) -> Tuple[List[TaskWrite], List[str]]:
    """
    Deja una sola fila por `dedupe_key`, quedándose con la última.

    No es una optimización: Postgres aborta con 21000 —y PostgREST devuelve un
    500— si un `ON CONFLICT DO UPDATE` toca la misma fila dos veces en la misma
    sentencia, así que un folio repetido dentro de una hoja tumba el lote
    entero. En el export real son dos casos, los dos en EDUARDO MANZANARES: uno
    es la misma fila copiada dos veces y el otro son dos actividades distintas
    a las que alguien puso el mismo folio.

    Se conserva la última porque en estos trackers las filas se añaden y se
    editan hacia abajo, así que la de más abajo es la versión más reciente. Las
    colapsadas se devuelven para reportarlas: el arreglo de fondo es corregir
    el folio en la hoja, y eso no lo puede hacer este script.
    """
    por_clave: Dict[str, TaskWrite] = {}
    orden: List[str] = []
    sueltas: List[TaskWrite] = []
    repetidas: List[str] = []
    for tarea in tareas:
        clave = compute_dedupe_key(tarea.folio, hoja)
        if not clave:
            sueltas.append(tarea)
            continue
        if clave in por_clave:
            repetidas.append(str(tarea.folio))
        else:
            orden.append(clave)
        por_clave[clave] = tarea
    return [por_clave[c] for c in orden] + sueltas, repetidas


def clasificar(
    repo: TaskRepository, hoja: str, tareas: Sequence[TaskWrite]
) -> Tuple[List[str], List[str], List[str]]:
    """
    `(altas, actualizaciones, sin_clave)` por `dedupe_key`.

    Es la cuenta que decide si el upsert va a insertar o a fusionar, y se hace
    con la misma función que usa la escritura (`compute_dedupe_key`), no con
    una regla paralela que pudiera discrepar.
    """
    claves: List[str] = []
    sin_clave: List[str] = []
    for tarea in tareas:
        clave = compute_dedupe_key(tarea.folio, hoja)
        if clave:
            claves.append(clave)
        else:
            sin_clave.append(str(tarea.folio))

    # En trozos: `por_dedupe_keys` monta un `in.(...)` en la URL, y las 1.883
    # claves de ADMINISTRADOR la pasan del límite del servidor (HTTP 400).
    existentes: Dict[str, Any] = {}
    for inicio in range(0, len(claves), LOTE_CONSULTA):
        existentes.update(repo.por_dedupe_keys(claves[inicio : inicio + LOTE_CONSULTA]))
    altas = [c for c in claves if c not in existentes]
    actualizaciones = [c for c in claves if c in existentes]
    return altas, actualizaciones, sin_clave, set(existentes)


def apartar_incompletas(
    hoja: str, tareas: Sequence[TaskWrite], existentes: Set[str]
) -> Tuple[List[TaskWrite], List[Tuple[str, str]]]:
    """
    Separa las altas a las que les falta una columna NOT NULL.

    El repositorio ya se niega a insertarlas —y hace bien: `concepto` es NOT
    NULL sin DEFAULT—, pero lo hace abortando el lote entero. Aquí se apartan
    antes para que las demás filas de la hoja sí entren, y se reportan con su
    folio para que alguien complete el dato en el tracker.

    Rellenarlas con un texto de relleno sería inventar contenido en una tabla
    de producción, así que no se hace.
    """
    cargables: List[TaskWrite] = []
    apartadas: List[Tuple[str, str]] = []
    for tarea in tareas:
        clave = compute_dedupe_key(tarea.folio, hoja)
        if clave and clave not in existentes:
            faltantes = falta_para_el_alta(tarea)
            if faltantes:
                apartadas.append((str(tarea.folio), ", ".join(faltantes)))
                continue
        cargables.append(tarea)
    return cargables, apartadas


def _cargar_ventas(ruta: pathlib.Path, engine: Any, settings: Any, args: Any) -> int:
    """
    Carga las hojas `(VENTAS)` en `quotes`.

    Camino paralelo al de tareas y por los mismos motivos: se reutiliza
    `QuoteRepository.guardar_lote`, que ya resuelve `vendedor_id`, las columnas
    NOT NULL y el upsert sobre `(folio, source_sheet)`. Aquí no hace falta
    colapsar folios repetidos —la clave incluye la hoja, así que dos personas
    con el mismo folio no colisionan— pero sí apartar las altas sin folio, que
    es la única obligatoria que el export puede no traer.
    """
    from backend.repositories.quotes import QuoteRepository

    repo = QuoteRepository(engine, settings)
    print(f"Archivo: {ruta.name}")
    hojas = leer_hojas_de_ventas(ruta, args.hoja)
    print(f"Hojas con firma de ventas: {len(hojas)}\n")

    total = collections.Counter()
    detalle: List[Tuple[str, int, int, int]] = []
    rechazos_por_hoja: Dict[str, List[Tuple[int, str]]] = {}

    for hoja, filas in sorted(hojas.items()):
        cotizaciones, rechazos = convertir_ventas(filas)
        if rechazos:
            rechazos_por_hoja[hoja] = rechazos

        folios = [str(c.folio).strip() for c in cotizaciones
                  if c.folio is not None and str(c.folio).strip()]
        existentes = {}
        for inicio in range(0, len(folios), LOTE_CONSULTA):
            existentes.update(repo.por_folios(folios[inicio : inicio + LOTE_CONSULTA]))
        altas = sum(1 for f in folios if (f, hoja) not in existentes)

        total["filas"] += len(filas)
        total["altas"] += altas
        total["actualizaciones"] += len(folios) - altas
        total["rechazos"] += len(rechazos)
        detalle.append((hoja, altas, len(folios) - altas, len(rechazos)))

        if args.ejecutar and cotizaciones:
            for inicio in range(0, len(cotizaciones), LOTE):
                repo.guardar_lote(hoja, cotizaciones[inicio : inicio + LOTE])

    ancho = max((len(h) for h, *_ in detalle), default=10)
    print(f"{'hoja'.ljust(ancho)}  {'altas':>6} {'actual.':>8} {'rechazo':>8}")
    for hoja, altas, actualizaciones, rechazos in detalle:
        print(f"{hoja.ljust(ancho)}  {altas:>6} {actualizaciones:>8} {rechazos:>8}")

    print(f"\nTOTAL  filas={total['filas']}  altas={total['altas']}  "
          f"actualizaciones={total['actualizaciones']}  rechazos={total['rechazos']}")

    for hoja, rechazos in rechazos_por_hoja.items():
        print(f"\nRECHAZOS en {hoja!r}:")
        for numero, motivo in rechazos[:10]:
            print(f"  fila {numero}: {motivo}")
        if len(rechazos) > 10:
            print(f"  ... y {len(rechazos) - 10} más")

    print("\n" + ("ESCRITO EN LA BASE." if args.ejecutar else "DRY-RUN: no se escribió nada."))
    return 0


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--archivo", required=True, help="Ruta del .xlsx exportado del Tracker")
    parser.add_argument("--hoja", default=None, help="Procesar solo esta hoja")
    parser.add_argument(
        "--ejecutar",
        action="store_true",
        help="Escribe de verdad en la base. Sin esta bandera solo se cuenta.",
    )
    parser.add_argument(
        "--ventas",
        action="store_true",
        help=("Carga las hojas `(VENTAS)` en `quotes` en vez de las de tareas en "
              "`tasks`. Son tablas distintas y nunca se mezclan."),
    )
    args = parser.parse_args()

    ruta = pathlib.Path(args.archivo)
    if not ruta.exists():
        print(f"No existe el archivo: {ruta}", file=sys.stderr)
        return 2

    _cargar_env()
    settings = cargar_settings()
    if not settings.hay_postgrest:
        print("Faltan SUPABASE_URL / SUPABASE_KEY.", file=sys.stderr)
        return 2
    engine = PostgrestEngine(settings.supabase_url, settings.supabase_key)

    if args.ventas:
        return _cargar_ventas(ruta, engine, settings, args)

    repo = TaskRepository(engine, settings)

    print(f"Archivo: {ruta.name}")
    hojas = leer_hojas(ruta, args.hoja)
    print(f"Hojas con firma de tareas: {len(hojas)}\n")

    total = collections.Counter()
    rechazos_por_hoja: Dict[str, List[Tuple[int, str]]] = {}
    incompletas_por_hoja: Dict[str, List[Tuple[str, str]]] = {}
    repetidas_por_hoja: Dict[str, List[str]] = {}
    detalle: List[Tuple[str, int, int, int]] = []

    for hoja, filas in sorted(hojas.items()):
        tareas, rechazos = convertir(filas)
        if rechazos:
            rechazos_por_hoja[hoja] = rechazos
        tareas, repetidas = colapsar_claves_repetidas(hoja, tareas)
        if repetidas:
            repetidas_por_hoja[hoja] = repetidas
        altas, actualizaciones, sin_clave, existentes = clasificar(repo, hoja, tareas)
        cargables, incompletas = apartar_incompletas(hoja, tareas, existentes)
        if incompletas:
            incompletas_por_hoja[hoja] = incompletas

        total["filas"] += len(filas)
        total["altas"] += len(altas) - len(incompletas)
        total["actualizaciones"] += len(actualizaciones)
        total["rechazos"] += len(rechazos)
        total["sin_clave"] += len(sin_clave)
        total["incompletas"] += len(incompletas)
        total["repetidas"] += len(repetidas)
        detalle.append(
            (hoja, len(altas) - len(incompletas), len(actualizaciones), len(rechazos))
        )

        if args.ejecutar and cargables:
            for inicio in range(0, len(cargables), LOTE):
                repo.guardar_lote(hoja, cargables[inicio : inicio + LOTE])

    ancho = max((len(h) for h, *_ in detalle), default=10)
    print(f"{'hoja'.ljust(ancho)}  {'altas':>6} {'actual.':>8} {'rechazo':>8}")
    for hoja, altas, actualizaciones, rechazos in detalle:
        if altas or rechazos:
            print(f"{hoja.ljust(ancho)}  {altas:>6} {actualizaciones:>8} {rechazos:>8}")

    print(
        f"\nTOTAL  filas={total['filas']}  altas={total['altas']}  "
        f"actualizaciones={total['actualizaciones']}  "
        f"rechazos={total['rechazos']}  sin_clave={total['sin_clave']}  "
        f"altas_incompletas={total['incompletas']}  "
        f"folios_repetidos={total['repetidas']}"
    )

    for hoja, repetidas in repetidas_por_hoja.items():
        print(f"\nFOLIOS REPETIDOS en {hoja!r} (se conservó la última fila): {', '.join(repetidas)}")

    for hoja, incompletas in incompletas_por_hoja.items():
        print(f"\nALTAS APARTADAS en {hoja!r} (falta una columna obligatoria):")
        for folio, faltantes in incompletas[:10]:
            print(f"  folio {folio}: sin {faltantes}")
        if len(incompletas) > 10:
            print(f"  ... y {len(incompletas) - 10} más")

    for hoja, rechazos in rechazos_por_hoja.items():
        print(f"\nRECHAZOS en {hoja!r}:")
        for numero, motivo in rechazos[:10]:
            print(f"  fila {numero}: {motivo}")
        if len(rechazos) > 10:
            print(f"  ... y {len(rechazos) - 10} más")

    print("\n" + ("ESCRITO EN LA BASE." if args.ejecutar else "DRY-RUN: no se escribió nada."))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
