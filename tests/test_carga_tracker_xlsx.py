"""
Pruebas del cargador de export .xlsx (`scripts/cargar_tracker_xlsx.py`).

Se prueban las funciones puras —localizar el encabezado, decidir qué hoja es de
tareas, y coaccionar los tipos que Excel inventa— porque son las tres que
deciden qué llega a la base. La escritura no se prueba aquí: es
`TaskRepository.guardar_lote`, que ya tiene sus propias pruebas.

Cada caso viene de una patología observada en el export real de 110 hojas que
motivó el script, no de una hipótesis: la fila de encabezado que se mueve entre
la 9 y la 12, las hojas de ventas que comparten `FOLIO`/`CONCEPTO` con las de
tareas, el folio timestamp que Excel entrega como `float`, la fecha escrita
dentro de un campo de texto libre, y la columna `RELOJ` que en el .xlsx viene
como serial de fecha y en la base es otra cosa.
"""

from __future__ import annotations

import datetime
import importlib.util
import pathlib

import openpyxl
import pytest

RAIZ = pathlib.Path(__file__).resolve().parent.parent


def _cargar_modulo():
    """El script vive en `scripts/`, que no es un paquete importable."""
    ruta = RAIZ / "scripts" / "cargar_tracker_xlsx.py"
    spec = importlib.util.spec_from_file_location("cargar_tracker_xlsx", ruta)
    modulo = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(modulo)
    return modulo


carga = _cargar_modulo()


def _hoja(encabezados, filas, fila_encabezado=10):
    """Hoja en memoria con el encabezado en la fila indicada."""
    libro = openpyxl.Workbook()
    hoja = libro.active
    for columna, texto in enumerate(encabezados, start=1):
        hoja.cell(fila_encabezado, columna, texto)
    for desplazamiento, fila in enumerate(filas, start=1):
        for columna, valor in enumerate(fila, start=1):
            hoja.cell(fila_encabezado + desplazamiento, columna, valor)
    return hoja


ENCABEZADO_TAREAS = [
    "Folio", "ALTA", "FECHA", "Hora", "Clasificacion", "Concepto", "Involucrados",
    "Avance %", "Fecha estimada de fin", "Hora estimada de fin", "Reloj",
    "Restricciones", "Prioridades", "Riesgos", "Fecha respuesta", "Correo",
    "Carpeta", "Status",
]

ENCABEZADO_VENTAS = [
    "FOLIO", "AREA", "CLIENTE", "CONCEPTO", "CLASIFICACION", "VENDEDOR",
    "F. VISITA", "F. INICIO", "F. ENTREGA", "DIAS", "AVANCE", "ESTATUS",
    "COMENTARIOS", "REQUISITOR", "PRIO. COT.", "INFO CLIENTE", "F2",
    "COTIZACION", "TIMELINE", "LAYOUT",
]


class TestLocalizarEncabezado:
    """La fila del encabezado no es fija en las hojas reales."""

    @pytest.mark.parametrize("fila", [9, 10, 12])
    def test_encuentra_el_encabezado_en_cualquiera_de_las_filas_reales(self, fila):
        hoja = _hoja(ENCABEZADO_TAREAS, [["RR-1", None]], fila_encabezado=fila)
        ubicacion = carga.localizar_encabezado(hoja)
        assert ubicacion is not None
        assert ubicacion[0] == fila

    def test_una_hoja_de_ventas_no_es_una_hoja_de_tareas(self):
        """
        Comparte `FOLIO`, `CONCEPTO` y `AVANCE` con las de tareas, pero alimenta
        `quotes` con otro esquema. Tratarla como tarea metería cotizaciones en
        `tasks`.
        """
        hoja = _hoja(ENCABEZADO_VENTAS, [["AV-1", "OBRA"]], fila_encabezado=1)
        assert carga.localizar_encabezado(hoja) is None

    def test_una_hoja_sin_avance_ni_status_no_es_de_tareas(self):
        """Las 16 hojas-plantilla del export traen solo un formulario vacío."""
        hoja = _hoja(
            ["Folio", "Fecha estimada de fin", "Involucrados", "Concepto", "Riesgos"],
            [["X-1", None, None, None, None]],
            fila_encabezado=2,
        )
        assert carga.localizar_encabezado(hoja) is None


class TestFilasDeHoja:
    def test_descarta_las_filas_sin_folio(self):
        """Las hojas reales llegan vacías hasta la fila 1000 y pico."""
        hoja = _hoja(ENCABEZADO_TAREAS, [["RR-1"], [None], [""], ["RR-2"]])
        filas = list(carga.filas_de_hoja(hoja, 10, ENCABEZADO_TAREAS))
        assert [f["Folio"] for f in filas] == ["RR-1", "RR-2"]

    def test_conserva_el_numero_de_fila_del_xlsx(self):
        """Sin él, un rechazo no se puede localizar en el archivo."""
        hoja = _hoja(ENCABEZADO_TAREAS, [["RR-1"]])
        assert list(carga.filas_de_hoja(hoja, 10, ENCABEZADO_TAREAS))[0]["_fila_xlsx"] == 11

    def test_no_incluye_la_columna_reloj(self):
        """
        `RELOJ` sale del .xlsx como serial de fecha y en la base es otra cosa.
        Omitirla deja que el `merge-duplicates` conserve el valor real.
        """
        hoja = _hoja(ENCABEZADO_TAREAS, [["RR-1"]])
        fila = list(carga.filas_de_hoja(hoja, 10, ENCABEZADO_TAREAS))[0]
        assert "Reloj" not in fila


class TestCoercionDeTipos:
    """Excel tipa la celda; la base tipa la columna. Aquí se concilian."""

    def test_el_folio_timestamp_conserva_el_punto_cero(self):
        """
        El `.0` no es un descuido: es lo que la base tiene guardado.

        Excel entrega estos folios como `float` y son 43 filas del export. Las
        112 filas con folio timestamp que ya viven en `tasks` están
        almacenadas como `'1772639658256.0'`, y su `dedupe_key` es esa misma
        cadena. Recortar el `.0` daría una clave distinta a la almacenada y el
        upsert insertaría un duplicado en vez de actualizar la fila.
        """
        assert carga._a_texto_si_toca("Folio", 1772639658256.0) == "1772639658256.0"

    def test_una_fecha_en_texto_libre_toma_la_forma_que_ya_usa_la_base(self):
        """
        `restricciones` es texto libre y hay quien escribe una fecha dentro. La
        base guarda catorce de esas filas como `"2026-07-17 00:00:00"`.
        """
        valor = carga._a_texto_si_toca("Restricciones", datetime.datetime(2026, 7, 17))
        assert valor == "2026-07-17 00:00:00"

    def test_el_avance_conserva_su_tipo_numerico(self):
        """
        La distinción entre `1` y `"1"` decide el archivado (AGENTS.md §4).
        Convertirlo a texto aquí la destruiría.
        """
        assert carga._a_texto_si_toca("Avance %", 100.0) == 100.0

    def test_una_fecha_en_columna_de_fecha_llega_intacta(self):
        """`a_fecha` necesita el `datetime`, no su `str()`."""
        fecha = datetime.datetime(2026, 7, 17)
        assert carga._a_texto_si_toca("FECHA", fecha) is fecha

    def test_un_encabezado_desconocido_pasa_sin_tocarse(self):
        """No es columna de `tasks`: `desde_hoja` lo descartará después."""
        assert carga._a_texto_si_toca("COLUMNA INVENTADA", 5.0) == 5.0


class TestConvertir:
    def test_una_fila_invalida_no_tumba_a_las_buenas(self):
        """
        Un export manual trae erratas puntuales. Perder las 300 filas buenas de
        un tracker por una celda mal escrita sería peor que apartar esa celda.
        """
        filas = [
            {"Folio": "RR-1", "Concepto": "Buena", "_fila_xlsx": 11},
            {"Folio": "RR-2", "Concepto": {"no": "texto"}, "_fila_xlsx": 12},
            {"Folio": "RR-3", "Concepto": "Otra buena", "_fila_xlsx": 13},
        ]
        convertidas, rechazos = carga.convertir(filas)
        assert [t.folio for t in convertidas] == ["RR-1", "RR-3"]
        assert [numero for numero, _ in rechazos] == [12]

    def test_una_hoja_sin_folios_repetidos_no_se_altera(self):
        tareas, _ = carga.convertir(
            [
                {"Folio": "RM-1", "Concepto": "A", "_fila_xlsx": 11},
                {"Folio": "RM-2", "Concepto": "B", "_fila_xlsx": 12},
            ]
        )
        colapsadas, repetidas = carga.colapsar_claves_repetidas("HOJA", tareas)
        assert [t.folio for t in colapsadas] == ["RM-1", "RM-2"]
        assert repetidas == []

    def test_un_folio_repetido_deja_una_fila_y_gana_la_ultima(self):
        """
        Postgres aborta con 21000 si el upsert toca la misma fila dos veces, así
        que un folio repetido en la hoja tumbaba el lote entero con un 500.
        """
        tareas, _ = carga.convertir(
            [
                {"Folio": "RM-0156", "Concepto": "PREVENTIVOS", "_fila_xlsx": 15},
                {"Folio": "RM-9", "Concepto": "Otra", "_fila_xlsx": 20},
                {"Folio": "RM-0156", "Concepto": "CORRECTIVO", "_fila_xlsx": 41},
            ]
        )
        colapsadas, repetidas = carga.colapsar_claves_repetidas("HOJA", tareas)
        assert [t.folio for t in colapsadas] == ["RM-0156", "RM-9"]
        assert [t.concepto for t in colapsadas] == ["CORRECTIVO", "Otra"]
        assert repetidas == ["RM-0156"]


class TestAltasIncompletas:
    """`concepto` es NOT NULL sin DEFAULT: un alta sin él aborta el lote."""

    def test_un_alta_sin_concepto_se_aparta(self):
        tareas, _ = carga.convertir([{"Folio": "RM-1", "_fila_xlsx": 11}])
        cargables, apartadas = carga.apartar_incompletas("HOJA", tareas, set())
        assert cargables == []
        assert apartadas == [("RM-1", "concepto")]

    def test_una_actualizacion_sin_concepto_si_pasa(self):
        """
        En una actualización el repositorio toma `concepto` de la fila que ya
        está en la base (`_completar_obligatorias`), así que no falta nada.
        """
        tareas, _ = carga.convertir([{"Folio": "RM-1", "_fila_xlsx": 11}])
        cargables, apartadas = carga.apartar_incompletas("HOJA", tareas, {"HOJA::RM-1"})
        assert [t.folio for t in cargables] == ["RM-1"]
        assert apartadas == []

    def test_un_alta_completa_pasa(self):
        tareas, _ = carga.convertir([{"Folio": "RM-1", "Concepto": "Algo", "_fila_xlsx": 11}])
        cargables, apartadas = carga.apartar_incompletas("HOJA", tareas, set())
        assert [t.folio for t in cargables] == ["RM-1"]
        assert apartadas == []

    def test_el_rechazo_dice_en_que_fila_del_xlsx_esta(self):
        filas = [{"Folio": "RR-2", "Concepto": {"no": "texto"}, "_fila_xlsx": 99}]
        _, rechazos = carga.convertir(filas)
        assert rechazos[0][0] == 99
        assert "concepto" in rechazos[0][1].lower()
