"""
Pruebas de `scripts/sincronizar_terminadas.py`.

El script concilia dos verdades sobre qué es una tarea «terminada»:

* En la **hoja** es una posición: arriba las activas, un hueco, y debajo la
  sección de realizadas. No hay rótulo; ninguna de las 40 hojas escribe
  «TAREAS REALIZADAS», solo dejan filas vacías.
* En la **base** es una propiedad calculada de la fila (`_esta_archivada`).

Discrepan en 210 filas que su dueño bajó a realizadas sin tocarles el AVANCE, y
en 959 copias de un mismo folio donde una está cerrada y la otra no. Cada caso
de aquí sale de una de esas dos patologías medidas, no de una hipótesis.
"""

from __future__ import annotations

import importlib.util
import pathlib

import openpyxl
import pytest

RAIZ = pathlib.Path(__file__).resolve().parent.parent


def _cargar_modulo():
    """El script vive en `scripts/`, que no es un paquete importable."""
    ruta = RAIZ / "scripts" / "sincronizar_terminadas.py"
    spec = importlib.util.spec_from_file_location("sincronizar_terminadas", ruta)
    modulo = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(modulo)
    return modulo


sync = _cargar_modulo()


def _hoja(folios_por_fila):
    """Hoja con encabezado en la fila 10 y folios en las filas indicadas."""
    libro = openpyxl.Workbook()
    hoja = libro.active
    hoja.cell(10, 1, "Folio")
    for fila, folio in folios_por_fila.items():
        hoja.cell(fila, 1, folio)
    return hoja


class TestSepararPorHueco:
    """El separador de la hoja es un hueco, no un rótulo."""

    def test_parte_en_el_primer_hueco_grande(self):
        """
        La hoja de Geraldine: nueve activas en 11-19, vacío hasta 274, y el
        histórico desde 275.
        """
        hoja = _hoja({**{r: f"GM-{r}" for r in range(11, 20)},
                      **{r: f"GM-{r}" for r in range(275, 280)}})
        arriba, abajo = sync.separar_por_hueco(hoja, 10)
        assert arriba == list(range(11, 20))
        assert abajo == list(range(275, 280))

    def test_un_renglon_suelto_en_blanco_no_parte_la_hoja(self):
        """
        Hay hojas con una o dos filas vacías entre bloques que no son el
        separador. Cortar ahí mandaría el bloque activo a realizadas.
        """
        hoja = _hoja({11: "A-1", 12: "A-2", 14: "A-3", 15: "A-4"})
        arriba, abajo = sync.separar_por_hueco(hoja, 10)
        assert arriba == [11, 12, 14, 15]
        assert abajo == []

    def test_una_hoja_sin_hueco_no_tiene_seccion_de_realizadas(self):
        hoja = _hoja({r: f"A-{r}" for r in range(11, 16)})
        arriba, abajo = sync.separar_por_hueco(hoja, 10)
        assert arriba == list(range(11, 16))
        assert abajo == []

    def test_una_hoja_vacia_no_revienta(self):
        assert sync.separar_por_hueco(_hoja({}), 10) == ([], [])


class TestRotulo:
    """El separador de verdad es el rótulo «TAREAS REALIZADAS»."""

    def _hoja_con_rotulo(self, fila_rotulo, columna_rotulo=6):
        libro = openpyxl.Workbook()
        hoja = libro.active
        hoja.cell(10, 1, "Folio")
        for r in range(11, fila_rotulo):
            hoja.cell(r, 1, f"A-{r}")
        hoja.cell(fila_rotulo, columna_rotulo, "TAREAS REALIZADAS")
        for r in range(fila_rotulo + 1, fila_rotulo + 4):
            hoja.cell(r, 1, f"B-{r}")
        return hoja

    def test_lo_encuentra_en_la_columna_del_concepto(self):
        """No está en la columna del folio sino en la del concepto."""
        assert sync.fila_del_rotulo(self._hoja_con_rotulo(41), 10) == 41

    def test_lo_encuentra_aunque_este_muy_abajo(self):
        """En ADMINISTRADOR está en la fila 1.864: acotar la búsqueda lo pierde."""
        assert sync.fila_del_rotulo(self._hoja_con_rotulo(1864), 10) == 1864

    def test_devuelve_none_si_la_hoja_no_lo_escribe(self):
        libro = openpyxl.Workbook()
        hoja = libro.active
        hoja.cell(10, 1, "Folio")
        hoja.cell(11, 1, "A-1")
        assert sync.fila_del_rotulo(hoja, 10) is None

    def test_no_confunde_un_concepto_que_dice_realizar(self):
        """`REALIZAR REPORTE` aparece en decenas de conceptos y no es el rótulo."""
        libro = openpyxl.Workbook()
        hoja = libro.active
        hoja.cell(10, 1, "Folio")
        hoja.cell(11, 6, "REALIZAR REPORTE DE CHECADAS")
        assert sync.fila_del_rotulo(hoja, 10) is None

    def test_el_rotulo_manda_sobre_el_hueco(self):
        """
        En `SONIA GARCIA PEREZ` el hueco deja 247 filas arriba y el rótulo solo
        29. Donde hay rótulo, el hueco se ignora.
        """
        libro = openpyxl.Workbook()
        hoja = libro.active
        hoja.cell(10, 1, "Folio")
        for r in range(11, 15):
            hoja.cell(r, 1, f"A-{r}")
        hoja.cell(15, 6, "TAREAS REALIZADAS")
        for r in range(16, 20):
            hoja.cell(r, 1, f"B-{r}")
        assert sync.filas_de_realizadas(hoja, 10) == [16, 17, 18, 19]

    def test_sin_rotulo_se_cae_al_hueco(self):
        """31 hojas no lo escriben y seguirían sin sección de realizadas."""
        hoja = _hoja({11: "A-1", 12: "A-2", 40: "B-1", 41: "B-2"})
        assert sync.filas_de_realizadas(hoja, 10) == [40, 41]


class TestDesajustesPorSeparador:
    def test_una_fila_bajo_el_separador_y_abierta_se_detecta(self):
        filas = [{"dedupe_key": "k1", "folio": "GM-1", "source_sheet": "HOJA",
                  "avance": 10.0, "status": "ASIGNADO", "cumplimiento": None}]
        pendientes = sync.desajustes_por_separador(filas, {"HOJA": {"GM-1": 10.0}})
        assert [p["folio"] for p in pendientes] == ["GM-1"]

    def test_una_fila_bajo_el_separador_pero_ya_cerrada_se_deja(self):
        filas = [{"dedupe_key": "k1", "folio": "GM-1", "source_sheet": "HOJA",
                  "avance": 100.0, "status": "ASIGNADO", "cumplimiento": None}]
        assert sync.desajustes_por_separador(filas, {"HOJA": {"GM-1": 100.0}}) == []

    def test_una_fila_de_arriba_no_se_toca(self):
        """Está en operativo en la hoja: cerrarla sería inventar el cierre."""
        filas = [{"dedupe_key": "k1", "folio": "GM-9", "source_sheet": "HOJA",
                  "avance": 10.0, "status": "ASIGNADO", "cumplimiento": None}]
        assert sync.desajustes_por_separador(filas, {"HOJA": {"GM-1": 10.0}}) == []

    def test_el_nombre_de_hoja_se_compara_normalizado(self):
        """
        `source_sheet` no está normalizado: hay hojas con espacio inicial
        (` LILIANA AYLIN MARTINEZ IBARRA`) y capitalización mixta.
        """
        filas = [{"dedupe_key": "k1", "folio": "GM-1", "source_sheet": " Hoja  Rara ",
                  "avance": 0.0, "status": "ASIGNADO", "cumplimiento": None}]
        assert len(sync.desajustes_por_separador(filas, {"HOJA RARA": {"GM-1": 0.0}})) == 1

    def test_converge_aunque_la_hoja_diga_otro_avance(self):
        """
        Estar debajo del rótulo cierra la fila al 100 %, así que la hoja seguirá
        diciendo 33 % para siempre. Si el desajuste se midiera comparando
        avances, esta fila se re-detectaría en cada pasada y el script no
        convergería nunca.
        """
        filas = [{"dedupe_key": "k1", "folio": "GM-1", "source_sheet": "HOJA",
                  "avance": 100.0, "status": "ASIGNADO", "cumplimiento": "SI"}]
        assert sync.desajustes_por_separador(filas, {"HOJA": {"GM-1": 33.0}}) == []


class TestCopiasDesincronizadas:
    def _fila(self, clave, hoja, avance, concepto="Actividad normal"):
        return {"dedupe_key": clave, "folio": "PPC-1", "source_sheet": hoja,
                "avance": avance, "status": "ASIGNADO", "cumplimiento": None,
                "concepto": concepto}

    def test_la_copia_abierta_toma_el_cierre_de_su_hermana(self):
        """
        El caso real: la fila del PPC maestro quedó en 0 % y la de la persona
        que hizo el trabajo en 100 %.
        """
        abierta = self._fila("PPC-1", "ADMINISTRADOR", 0.0)
        cerrada = self._fila("ZAIRA::PPC-1", "ZAIRA", 100.0)
        pendientes = sync.copias_desincronizadas([abierta, cerrada])
        assert pendientes == [(abierta, cerrada)]

    def test_sin_hermana_cerrada_no_se_cierra_nada(self):
        """Si todas están abiertas, inventar un cierre sería falsear el dato."""
        filas = [self._fila("PPC-1", "A", 10.0), self._fila("B::PPC-1", "B", 20.0)]
        assert sync.copias_desincronizadas(filas) == []

    def test_un_folio_con_una_sola_fila_no_es_una_copia(self):
        assert sync.copias_desincronizadas([self._fila("PPC-1", "A", 0.0)]) == []

    def test_gana_la_hermana_mas_avanzada(self):
        abierta = self._fila("PPC-1", "A", 0.0)
        filas = [abierta,
                 self._fila("B::PPC-1", "B", 100.0),
                 self._fila("C::PPC-1", "C", 100.0)]
        (_, modelo), = sync.copias_desincronizadas(filas)
        assert float(modelo["avance"]) == 100.0

    def test_una_microtarea_de_papa_caliente_nunca_propaga(self):
        """
        Comparten folio entre sí y con la cotización: propagar pondría el 100 %
        de una fase encima del trabajo pendiente de otra persona. Misma
        frontera que traza `campos_sincronizables`.
        """
        abierta = self._fila("PPC-1", "A", 0.0, concepto="Cotizar nave [Calculo y Diseño]")
        cerrada = self._fila("B::PPC-1", "B", 100.0)
        assert sync.copias_desincronizadas([abierta, cerrada]) == []


class TestFilaDeCierre:
    BASE = {"dedupe_key": "k1", "folio": "GM-1", "concepto": "Algo",
            "source_sheet": "HOJA", "avance": 10.0, "status": "ASIGNADO",
            "cumplimiento": None}

    def test_sin_modelo_cierra_con_avance_100_y_la_bandera(self):
        """
        Estar debajo del rótulo significa terminada, y un 10 % o 33 % ahí abajo
        es residuo de la plataforma anterior. Decisión del dueño del dato.
        `CUMPLIMIENTO` va igualmente porque es la bandera que no depende de
        cómo se lea el número.
        """
        carga = sync.fila_de_cierre(dict(self.BASE))
        assert carga["avance"] == 100
        assert carga["cumplimiento"] == "SI"


class TestCierreDeCotizacion:
    def _fila(self, estatus):
        return {"folio": "AV-1", "source_sheet": "ANTONIA_VENTAS",
                "avance": 0.0, "estatus": "ASIGNADO",
                "_hoja": {"avance": 0.0, "estatus": estatus}}

    def test_completada_es_boolean_y_no_el_texto_si(self):
        """
        `quotes.completada` es boolean en el esquema: mandarle `"SI"` aborta el
        lote con un 22P02 («invalid input syntax for type boolean»).
        """
        carga = sync._cierre_de_cotizacion(self._fila(None))
        assert carga["completada"] is True

    def test_cierra_con_avance_100(self):
        assert sync._cierre_de_cotizacion(self._fila(None))["avance"] == 100

    def test_conserva_el_estatus_terminal_de_la_hoja(self):
        """«Perdida x Tiempo» no es lo mismo que entregada, y eso alimenta KPIs."""
        carga = sync._cierre_de_cotizacion(self._fila("Perdida x Tiempo"))
        assert carga["estatus"] == "Perdida x Tiempo"

    def test_no_inventa_estatus_cuando_la_hoja_no_lo_trae(self):
        assert "estatus" not in sync._cierre_de_cotizacion(self._fila(None))


class TestFilaDeCierreConModelo:
    BASE = TestFilaDeCierre.BASE

    def test_con_modelo_copia_el_estado_de_la_hermana(self):
        modelo = {"avance": 100.0, "status": "TERMINADA", "cumplimiento": "SI"}
        carga = sync.fila_de_cierre(dict(self.BASE), modelo)
        assert carga["avance"] == 100.0
        assert carga["status"] == "TERMINADA"

    def test_no_copia_los_campos_vacios_del_modelo(self):
        """Mandar `None` a una columna NOT NULL aborta el lote con un 23502."""
        modelo = {"avance": 100.0, "status": "", "cumplimiento": None}
        carga = sync.fila_de_cierre(dict(self.BASE), modelo)
        assert "status" not in carga
        assert "cumplimiento" not in carga

    @pytest.mark.parametrize("columna", ["folio", "concepto", "source_sheet"])
    def test_reenvia_las_obligatorias(self, columna):
        """
        Un upsert valida las NOT NULL sobre la tupla del INSERT **antes** de
        resolver el conflicto, así que hay que mandarlas aunque no cambien.
        """
        assert columna in sync.fila_de_cierre(dict(self.BASE))

    def test_la_identidad_siempre_viaja(self):
        assert sync.fila_de_cierre(dict(self.BASE))["dedupe_key"] == "k1"
